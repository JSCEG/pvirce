from __future__ import annotations

import argparse
import csv
import json
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parent
DEFAULT_EXCEL = ROOT / r"Insumos/Actualizaciones 2Convoctaoria/Actualizacion17072026.xlsx"
DEFAULT_BASE_CSV = ROOT / "tmp_sheet_current.csv"
DEFAULT_OUTPUT_JSON = ROOT / "output/minuta_70py_payload.json"
DEFAULT_OUTPUT_CSV = ROOT / "output/minuta_70py_rows.csv"

MINUTA_COLUMNS = [
    "timestamp",
    "fecha_reunion",
    "gcr",
    "reunion_id",
    "reunion_nombre",
    "folio",
    "proyecto_id",
    "fila_sheet",
    "proyecto",
    "empresa_gie",
    "mw",
    "asistentes",
    "comentarios",
    "decision",
    "motivo_no_continua",
    "capturado_por",
]


@dataclass(frozen=True)
class BaseProjectRow:
    sheet_row_number: int
    project_index: int
    folio: str
    project_id: str
    project: str
    gie: str
    mw: str
    gcr: str


def clean_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("\xa0", " ").strip()


def parse_base_sheet(csv_path: Path) -> list[BaseProjectRow]:
    with csv_path.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.reader(handle))

    header = rows[7]
    index = {name: position for position, name in enumerate(header)}

    def col(row: list[str], name: str) -> str:
        position = index.get(name)
        if position is None or position >= len(row):
            return ""
        return clean_text(row[position])

    projects: list[BaseProjectRow] = []
    for full_index, row in enumerate(rows[8:]):
        if len(row) <= index.get("Pre Folio", -1):
            continue
        if not col(row, "Pre Folio").startswith("PRE-VUPE-C2-"):
            continue

        folio = col(row, "Folio Proyecto")
        if not folio or folio == "SIN INFORMACIÓN":
            continue

        mw = col(row, "Capacidad Contratada")
        if not mw or mw == "SIN INFORMACIÓN":
            mw = col(row, "Capacidad Generación")
        if not mw or mw == "SIN INFORMACIÓN":
            mw = col(row, "Capacidad Instalada Neta")

        projects.append(
            BaseProjectRow(
                sheet_row_number=9 + full_index,
                project_index=full_index,
                folio=folio,
                project_id=f"p_{full_index}",
                project=col(row, "Proyecto"),
                gie=col(row, "Grupo de Interés"),
                mw=mw,
                gcr=col(row, "Gerencia Regional"),
            )
        )

    return projects


def load_selected_folios(excel_path: Path, sheet_name: str) -> set[str]:
    workbook = openpyxl.load_workbook(excel_path, data_only=True)
    sheet = workbook[sheet_name]
    selected: set[str] = set()
    for row in range(2, sheet.max_row + 1):
        value = clean_text(sheet.cell(row, 1).value)
        if value:
            selected.add(value)
    return selected


def build_rows(
    base_projects: list[BaseProjectRow],
    selected_folios: set[str],
    reunion_id: str,
    reunion_nombre: str,
    fecha_reunion: str,
    capturado_por: str,
    asistentes: str,
    timestamp_base: str,
) -> list[dict[str, object]]:
    base_dt = datetime.fromisoformat(timestamp_base.replace("Z", "+00:00"))
    rows: list[dict[str, object]] = []

    for offset, project in enumerate(base_projects):
        is_selected = project.folio in selected_folios
        rows.append(
            {
                "timestamp": (base_dt + timedelta(milliseconds=offset)).isoformat().replace("+00:00", "Z"),
                "fecha_reunion": fecha_reunion,
                "gcr": project.gcr,
                "reunion_id": reunion_id,
                "reunion_nombre": reunion_nombre,
                "folio": project.folio,
                "proyecto_id": project.project_id,
                "fila_sheet": project.sheet_row_number,
                "proyecto": project.project,
                "empresa_gie": project.gie,
                "mw": project.mw,
                "asistentes": asistentes,
                "comentarios": "",
                "decision": "Continúa" if is_selected else "No continúa",
                "motivo_no_continua": "" if is_selected else "No seleccionado en la conciliación CFE-SENER",
                "capturado_por": capturado_por,
            }
        )

    return rows


def write_csv(path: Path, rows: list[dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=MINUTA_COLUMNS)
        writer.writeheader()
        writer.writerows(rows)


def main() -> int:
    parser = argparse.ArgumentParser(description="Build a Google Sheets minuta payload from Actualizacion17072026.xlsx")
    parser.add_argument("--excel", type=Path, default=DEFAULT_EXCEL)
    parser.add_argument("--base-csv", type=Path, default=DEFAULT_BASE_CSV)
    parser.add_argument("--sheet", default="70Py")
    parser.add_argument("--reunion-id", default="2026-07-17|NACIONAL|Reunión Selección Segunda Convocatoria CFE y SENER")
    parser.add_argument("--reunion-nombre", default="Reunión Selección Segunda Convocatoria CFE y SENER")
    parser.add_argument("--fecha-reunion", default="2026-07-17T07:00:00.000Z")
    parser.add_argument("--capturado-por", default="DGMESNIE")
    parser.add_argument(
        "--asistentes",
        default="Claudia Marcela Meza Vega <cmeza@energia.gob.mx>; Raúl Adame Ortiz <radame@energia.gob.mx>; Daniel Hilario Flores Vargas <dflores@energia.gob.mx>; Marian Olvera Lucas; Antonino López Ríos <antonino.lopez@cenace.gob.mx>",
    )
    parser.add_argument("--output-json", type=Path, default=DEFAULT_OUTPUT_JSON)
    parser.add_argument("--output-csv", type=Path, default=DEFAULT_OUTPUT_CSV)
    parser.add_argument("--timestamp-base", default="2026-07-17T00:00:00.000Z")
    args = parser.parse_args()

    selected_folios = load_selected_folios(args.excel, args.sheet)
    base_projects = parse_base_sheet(args.base_csv)
    rows = build_rows(
        base_projects=base_projects,
        selected_folios=selected_folios,
        reunion_id=args.reunion_id,
        reunion_nombre=args.reunion_nombre,
        fecha_reunion=args.fecha_reunion,
        capturado_por=args.capturado_por,
        asistentes=args.asistentes,
        timestamp_base=args.timestamp_base,
    )

    args.output_json.parent.mkdir(parents=True, exist_ok=True)
    args.output_json.write_text(json.dumps({"rows": rows}, ensure_ascii=False, indent=2), encoding="utf-8")
    write_csv(args.output_csv, rows)

    selected_count = sum(1 for row in rows if row["decision"] == "Continúa")
    print(f"rows={len(rows)} selected={selected_count} no_seleccionados={len(rows) - selected_count}")
    print(f"json={args.output_json}")
    print(f"csv={args.output_csv}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())